from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import redirect, render
from openpyxl import load_workbook
from django.db.models import Sum,Avg
from datetime import datetime


from .models import *
from .form import *
from .filter import BookFilter


def home(request):
    types = Book_type.objects.all()

    catagories = []
    datapoints = []
    for type in types:
        
        books = Books.objects.filter(category = type)
        expense = books.aggregate(Sum("distribution_expense"))["distribution_expense__sum"]
        average_expense = books.aggregate(average_expense = Avg("distribution_expense"))['average_expense']
        
        category = {"name":type,"books":len(books),"expense":expense,"average" : average_expense}
        catagories.append(category)
        y = round(expense) if len(books)>0 else 0
        data = {"label":type.type_name,"y":y}
        datapoints.append(data)
        
    all_books = Books.objects.all()
    total = {"books":len(all_books),"category":len(catagories),"expense":all_books.aggregate(sum = Sum('distribution_expense'))['sum'],"average":all_books.aggregate(avg = Avg('distribution_expense'))['avg']}
        
    context ={"datapoints":datapoints,"catagories":catagories,"total":total}
    
    return render(request,"home.html", context)
# book
def book_management(request):
    
    books_list = Books.objects.all()  
    bookFilter = BookFilter(request.GET, queryset=books_list)
    books_list = bookFilter.qs
    book_per_page = 15
    paginator = Paginator(books_list,book_per_page)
    page_num = request.GET.get('page',1)
    try:
        books = paginator.page(page_num)
        page_offset = (int(page_num)-1)*book_per_page 
    except PageNotAnInteger:
        # if page is not an integer, deliver the first page
        books = paginator.page(1)
        page_offset = 0
        
    except EmptyPage:
        # if the page is out of range, deliver the last page
        books = paginator.page(paginator.num_pages)
        page_offset = ((paginator.num_pages)-1)*book_per_page

    
    return render(request,'book_management.html',{"books":books,"bookFilter":bookFilter,"page_offset":page_offset})
def add_book(request):
    form = Books_form()
    message = ""
    # add individually 
    if request.method == 'POST' and 'idNumber' in request.POST:
        form = Books_form(request.POST) 

        if form and  form.is_valid():
            book =  form.save( commit=False)
            book.modification_log = "today |yosef |add"
            book.save()
            message = "Done,Record saved "
            form = Books_form()
        else:
            form = Books_form(request.POST)
            message = "unable to save please try again !"



    # import from excel sheet
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                id,title	,subtitle	,authors	,publisher,	published_date,	category,	distribution_expense = row
                category_typ = Book_type.objects.filter(type_name__exact = category.strip()).first()
                if category_typ == None:
                    modification_log = datetime.now().strftime("%Y-%m-%d %H:%M:%S") +" | "+str(request.user)+" | Imported"
                    Book_type.objects.create(type_name = category.strip(),modification_log = modification_log)
                modification_log = datetime.now().strftime("%Y-%m-%d %H:%M:%S") +" | "+str(request.user)+" | Imported"
                Books.objects.create( idNumber =id,title =title	, subtitle =subtitle	,author =authors	, publisher =publisher,	published_date =published_date,	category=category_typ,	 distribution_expense =distribution_expense ,modification_log = modification_log)
            
            context = {"header":"Successfully Imported",
                        "message":"Imported Books:"}
            return render(request, 'success.html',context)
        except Exception as  e:
                context = {"header":"Unable to Import !!!",
                           "message":e}
                return render(request, 'success.html',context)
      
    context = {
        "title":"Add Book",
        "header":"Add Book | Adding a Record ",
        "description":"Add a book, fill out all the required field, and save. You can add multiple records as you wish, and you have the option to import records from an Excel sheet by clicking the import button.",
        "submitBtn":"Save,Add Another",
        "form":form,
        "message":message,
        "add_category":add_category(request),
      
    }
    return render(request,'book_crud.html',context)
def editBook(request,id):
    book = Books.objects.get(pk = id)
    message = ""
    form = Books_form( instance=book)
    if request.method == 'POST':
        form = Books_form(request.POST, instance=book)
        if form.is_valid():
            book =  form.save( commit=False)
            book.modification_log ="negem ==="+ book.modification_log
            book.save()
            message = "Done,Record saved "
                    
            return redirect('books')  # Redirect to a success URL

        else:
            form = Books_form(instance=book)
            message = "unable to save please try again !"
    context = {
        "title":"Update Records",
        "header":"Update Records | Editing Books Data ",
        "description":"Edit book data effortlessly: simply modify the field data, save your changes, and watch as they're promptly updated in the database",
        "submitBtn":"Update",
        "form":form,
        "message":message,
         "add_category":add_category(request),
      
    }
    return render(request,'book_crud.html',context)
def deleteBook(request,id):
    book = Books.objects.get(pk = id)
    form = Books_form(instance=book)
    message = ""
    if request.method == 'POST':
        form = Books_form(request.POST, instance=book)

        try:
            book.delete()
            message = "Done,Record Deleted "
            return render(request,'success.html',{"message":message,"header":"Completed",'backto':'books'})
        except Exception as e:         
            message = "unable to save please try again !"+e

    context = {
        "title":"Delete Records",
        "header":"Delete | Erasing Books Data ",
        "description":"Erasing Books Data,Are you sure you want to delete this book? Deleting it is permanent, with no option for recovery once it's lost. If you're uncertain, please cancel the action",
        "submitBtn":"Delete ",
        "form":form,
        "message":message
      
    }
    return render(request,'book_crud.html',context)

# book category
def add_category(request):
    category_form = Category_form()
    if request.method == 'POST':
        category_form = Category_form(request.POST)
        if category_form.is_valid() :
            category = category_form.save(commit=False)
            category.modification_log = datetime.now().strftime("%Y-%m-%d %H:%M:%S") +" | "+str(request.user)+" | Create"
            category.save()
            category_form = Category_form()

    return category_form


# user
def user_managment(request):
    users = User.objects.all()

    return render(request,'user_managment.html',{'users':users })

def add_user(request): 
    message = "" 
    user_form = User_form()
    profile_form = Profile_form()
    if request.method == 'POST':
        user_form = User_form(request.POST)
        profile_form = Profile_form(request.POST,request.FILES)
        if user_form.is_valid() and profile_form.is_valid():
            print("-----------------")
            user =  user_form.save()
            user_profile = profile_form.save(commit=False)
            user_profile.user = user
            user_profile.save()
            user_form = User_form()
            profile_form = Profile_form()
            message = {'msg':"Done,User Created",'type':'success'}
        else:
            message = {'msg': "Unable to create user",'type':'danger'}

           
           
           
    context = {
        "title":"Add User",
        "header":"Add record | Creating New User ",
        "description":"Add a User, fill out all the required field, and save. You can add multiple records as you wish,",
        "submitBtn":"Save ",
        'user_form':user_form,
        'profile_form':profile_form,
        "message":message,
      
    }
            
    return render(request,'user_crud.html',context)
def editUser(request,id):
    user = User.objects.get(id = id)
    message = "" 
    
    if request.method == 'POST':
        user_form = User_form(request.POST)
        profile_form = Profile_form(request.POST,request.FILES)
        if user_form.is_valid() and profile_form.is_valid():
            user =  user_form.save()
            user_profile = profile_form.save(commit=False)
            user_profile.user = user
            user_profile.save()
            user_form = User_form()
            profile_form = Profile_form()
            message = {'msg':"Done,User Created",'type':'success'}
        else:
            message = {'msg': "Unable to create user",'type':'danger'}
    else:
        profile_form = Profile_form( instance=user.profile)

        user_form = User_form(instance = user)
           
           
           
    context = {
        "title":"Edit User",
        "header":"Edit record |Update User ",
        "description":"Edit User data effortlessly: simply modify the field data, save your changes, and watch as they're promptly updated in the database",
        "submitBtn":"Update ",
        'user_form':user_form,
        'profile_form':profile_form,
        "message":message,
      
    }
    return render(request,'user_crud.html',context)

    
def deleteUser(request,id):
    message = "" 
    user = User.objects.get(id = id)
    if request.method == 'POST':
        profile_form = Profile_form(request.POST,request.FILES,instance = user.profile)
        try:
            User.objects.get(id = id).delete()            
            message = "Done,User Account Removed"
            return render(request,'success.html',{"message":message,"header":"Completed",'backto':'users'})
            
        except Exception as a:
            message = {'msg': a,'type':'danger'}
            
    else:
        profile_form = Profile_form(instance=user.profile)
                      
    context = {
        "title":"Remove User",
        "header":"Data removing | Removing User Account ",
        "description":"Erasing User Data,Are you sure you want to delete this User? Deleting it is permanent, with no option for recovery once it's lost. If you're uncertain, please cancel the action",
        "submitBtn":"Delete ",
       
        'profile_form':profile_form,
        "message":message,
      
    }
            
    return render(request,'user_crud.html',context)
   